"""Turn the Excel time-series archive into tidy long series.

Every file in the archive shares one shape, which is what makes 981 files
tractable without 981 parsers:

    A                    B                 C      D      E      ...
1   Banking Survey: ...  المسح المصرفي     2017                  2018     <- year row, MERGED
2   (LE mn)              (بالمليون جنيه)   Jul.   Aug.   Sept.   Jan.     <- period row
3   Total Deposits       إجمالى الودائع    2940   3003   3050    3254     <- data
4                                                                          <- blank separator
5     Government Dep.    الودائع الحكومية  582    575    621     602      <- indent = child

Three things make it work:

  * The year row is merged across the months belonging to that calendar year.
    Unmerging and filling gives every column its year, which is the only way
    to know that "Jan." means the year *after* the one printed on the left --
    Egypt's fiscal year runs July to June.
  * Labels are bilingual, English in one column and Arabic in the other. Which
    column is which is detected by script, not assumed by position.
  * Indentation in the label encodes hierarchy, so it is preserved as a level
    rather than stripped.

Values here are period-end stocks (deposits *as of* June), so periods are
dated to the end of their window and `period_basis` says so.

Run:  python ingest/parse_excel.py [--limit N] [--category "GDP"]
"""

from __future__ import annotations

import argparse
import collections
import csv
import json
import os
import pathlib
import re
import sys

import openpyxl

ROOT = pathlib.Path(__file__).resolve().parent.parent
EXCEL = ROOT / "data" / "raw" / "excel"
MANIFEST = ROOT / "catalog" / "excel_manifest.json"
OUT_SERIES = ROOT / "data" / "clean" / "series" / "excel_archive.csv"
OUT_CATALOG = ROOT / "catalog" / "series.json"
QUARANTINE = ROOT / "catalog" / "excel_quarantine.json"

MONTHS = {
    "jan": 1, "january": 1, "feb": 2, "february": 2, "mar": 3, "march": 3,
    "apr": 4, "april": 4, "may": 5, "jun": 6, "june": 6, "jul": 7, "july": 7,
    "aug": 8, "august": 8, "sep": 9, "sept": 9, "september": 9,
    "oct": 10, "october": 10, "nov": 11, "november": 11, "dec": 12, "december": 12,
}
LAST_DAY = {1: 31, 2: 28, 3: 31, 4: 30, 5: 31, 6: 30,
            7: 31, 8: 31, 9: 30, 10: 31, 11: 30, 12: 31}

ARABIC = re.compile(r"[؀-ۿ]")
ARABIC_INDIC = str.maketrans("٠١٢٣٤٥٦٧٨٩", "0123456789")

# Short codes keep series IDs readable. Anything not listed falls back to a
# slug of the category name, so a new category still produces valid IDs.
CATEGORY_CODES = {
    "CBE": "CBE", "Domestic Debt": "DDEBT", "External Debt": "XDEBT",
    "GDP": "GDP", "Interest Rates": "IRATE", "Investments": "INV",
    "Net Foreign Direct Investment": "FDI", "Tourism": "TOUR",
    "Foreign Trade": "TRADE", "Stocks": "STOCK", "Banking Surveys": "BANK",
    "BOP": "BOP", "Inflation": "PRICE", "State Budget": "BUDGET",
}


class ParseError(RuntimeError):
    pass


def slug(text: str, limit: int = 44) -> str:
    s = re.sub(r"[^A-Za-z0-9]+", "_", (text or "")).strip("_").upper()
    s = re.sub(r"_+", "_", s)
    return s[:limit].strip("_")


def month_of(value) -> int | None:
    if value is None:
        return None
    token = re.sub(r"[^A-Za-z]", "", str(value)).lower()
    return MONTHS.get(token)


def year_of(value) -> int | None:
    if value is None:
        return None
    m = re.search(r"(19|20)\d{2}", str(value).translate(ARABIC_INDIC))
    return int(m.group(0)) if m else None


def to_number(value) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).translate(ARABIC_INDIC).strip()
    s = s.replace(",", "").replace("%", "").replace(" ", "")
    if s in {"", "-", "--", "..", "...", "N/A", "n/a", "NA", "*"}:
        return None
    negative = s.startswith("(") and s.endswith(")")  # accounting negatives
    if negative:
        s = s[1:-1]
    try:
        v = float(s)
    except ValueError:
        return None
    return -v if negative else v


# Some sheets report max_row as 65536 -- Excel's old limit -- because of stray
# formatting far below the data, and carry thousands of merged ranges down
# there with it. Building that grid takes minutes per file for nothing. The
# largest genuine table in the archive is 347 rows.
MAX_ROWS = 1500
MAX_COLS = 120
# A merge spanning this many cells is formatting, not a header spanning its
# months, so filling it would overwrite real values.
MAX_MERGE_CELLS = 400


def load_grid_xls(path: pathlib.Path) -> list[list]:
    """Legacy binary .xls, which openpyxl cannot read at all.

    A handful of the older files in the archive are genuinely BIFF rather than
    xlsx-with-the-wrong-extension, so they get xlrd. Merged ranges are filled
    the same way.
    """
    import xlrd

    book = xlrd.open_workbook(path, formatting_info=False)
    sheet = book.sheet_by_index(0)
    nrows = min(sheet.nrows, MAX_ROWS)
    ncols = min(sheet.ncols, MAX_COLS)
    rows = [
        [sheet.cell_value(r, c) if sheet.cell_value(r, c) != "" else None
         for c in range(ncols)]
        for r in range(nrows)
    ]
    while rows and not any(v is not None and str(v).strip() for v in rows[-1]):
        rows.pop()
    return rows


def load_grid(path: pathlib.Path) -> list[list]:
    """Sheet as a list of rows, with merged ranges filled in.

    Filling the merge is the whole trick: the year sits in one merged cell
    spanning six month columns, so without this every month but the first has
    no year attached.
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    ws = wb.active
    nrows = min(ws.max_row or 1, MAX_ROWS)
    ncols = min(ws.max_column or 1, MAX_COLS)
    rows = [
        [ws.cell(r, c).value for c in range(1, ncols + 1)]
        for r in range(1, nrows + 1)
    ]

    for rng in ws.merged_cells.ranges:
        if rng.min_row > nrows or rng.min_col > ncols:
            continue
        height = min(rng.max_row, nrows) - rng.min_row + 1
        width = min(rng.max_col, ncols) - rng.min_col + 1
        if height * width > MAX_MERGE_CELLS:
            continue
        value = ws.cell(rng.min_row, rng.min_col).value
        if value is None:
            continue
        for r in range(rng.min_row, rng.min_row + height):
            for c in range(rng.min_col, rng.min_col + width):
                rows[r - 1][c - 1] = value
    wb.close()

    # Drop the trailing empty rows the padding leaves behind.
    while rows and not any(v is not None and str(v).strip() for v in rows[-1]):
        rows.pop()
    return rows


# Balance-of-payments files skip the month row entirely and put fiscal years
# straight in the header: "2004/2005", "2005/06". Both widths occur.
_RE_FY = re.compile(r"\b((?:19|20)\d{2})\s*/\s*((?:19|20)?\d{2})\b")


def fy_of(value) -> tuple[int, str] | None:
    """-> (fiscal year end year, canonical label) for a header like 2005/06."""
    if value is None:
        return None
    m = _RE_FY.search(str(value).translate(ARABIC_INDIC))
    if not m:
        return None
    start = int(m.group(1))
    tail = m.group(2)
    end = int(tail) if len(tail) == 4 else start + 1
    if end < start or end - start > 1:
        return None
    return end, f"FY{start}/{str(end)[2:]}"


_RE_QUARTER = re.compile(r"^\s*Q\s*([1-4])\s*$", re.I)


def quarter_of(value) -> int | None:
    if value is None:
        return None
    m = _RE_QUARTER.match(str(value).translate(ARABIC_INDIC))
    return int(m.group(1)) if m else None


def find_period_row(grid: list[list]) -> tuple[int, str]:
    """-> (row index, strategy) for the row holding the period labels.

    The archive uses three header layouts and all three occur in volume:

      month        month names under a merged year row (most files)
      quarter      literal Q1..Q4, year merged above or embedded below
      fiscal_year  "2004/2005" straight in the header (balance of payments)

    Whichever row scores highest wins, so a new layout shows up as a
    quarantined file with its scores rather than as silently wrong data.
    """
    detectors = (
        ("month", month_of),
        ("quarter", quarter_of),
        ("fiscal_year", fy_of),
        ("year", year_of),
    )
    scores = {name: (-1, 0) for name, _ in detectors}
    for i, row in enumerate(grid[:14]):
        for name, fn in detectors:
            hits = sum(1 for cell in row if fn(cell))
            if hits > scores[name][1]:
                scores[name] = (i, hits)

    # Order matters. Quarter and fiscal-year headers are unambiguous. Month
    # names also appear inside quarter labels ("July/Sep 2023"), so they lose
    # ties. A bare year row is the weakest signal and goes last, because in a
    # monthly file it is the row *above* the real header.
    order = ("quarter", "fiscal_year", "month", "year")
    # Two columns is the normal case, but some files hold a single quarter, so
    # fall back to accepting one rather than quarantining real data.
    for threshold in (2, 1):
        for name in order:
            row, hits = scores[name]
            if hits >= threshold:
                return row, name
    raise ParseError(
        "no period row found (best scores: "
        + ", ".join(f"{k}={v[1]}" for k, v in scores.items())
        + ")"
    )


def build_periods(
    grid: list[list], period_row: int, frequency: str, strategy: str = "month"
) -> dict[int, dict]:
    """-> {column index: {period, fy, basis}}

    The year comes from the nearest row above that carries years; because
    merges have been filled, each column already has the right one.
    """
    if strategy == "fiscal_year":
        out: dict[int, dict] = {}
        for c, value in enumerate(grid[period_row]):
            parsed = fy_of(value)
            if parsed:
                end_year, label = parsed
                # Egypt's fiscal year closes 30 June.
                out[c] = {"period": f"{end_year}-06-30", "fy": label, "basis": "end"}
        if not out:
            raise ParseError("fiscal-year row had no parseable labels")
        return out

    if strategy == "year":
        # A bare year row, e.g. "End of June | 2005 | 2006 | ...". The month it
        # refers to is stated somewhere in the sheet's label column, so look
        # for it rather than assuming; Egypt's fiscal year ends in June, which
        # is the sensible default for annual data when nothing says otherwise.
        hint = " ".join(
            str(cell).lower()
            for row in grid[: period_row + 2]
            for cell in row[:3]
            if isinstance(cell, str)
        )
        month = 6 if "june" in hint else 12 if "december" in hint else (6 if frequency == "Annual" else 12)
        out = {}
        for c, value in enumerate(grid[period_row]):
            year = year_of(value)
            if year is None:
                continue
            fy = f"FY{year - 1}/{str(year)[2:]}" if month == 6 else str(year)
            out[c] = {"period": f"{year}-{month:02d}-{LAST_DAY[month]:02d}", "fy": fy, "basis": "end"}
        if not out:
            raise ParseError("year row had no parseable years")
        return out

    if strategy == "quarter":
        quarters = {c: q for c, q in
                    ((c, quarter_of(v)) for c, v in enumerate(grid[period_row])) if q}
        if not quarters:
            raise ParseError("quarter row had no parseable labels")

        # The year sits either in a merged row above (CBE files) or embedded in
        # the range printed below ("July/Sep 2023", balance of payments). Look
        # both ways rather than assuming, then carry the last known year
        # forward across columns that only inherit it.
        years: dict[int, int] = {}
        candidates = [period_row + 1, period_row + 2] + list(range(period_row - 1, max(-1, period_row - 4), -1))
        for c in quarters:
            for r in candidates:
                if 0 <= r < len(grid) and c < len(grid[r]):
                    y = year_of(grid[r][c])
                    if y:
                        years[c] = y
                        break

        # Egypt's fiscal quarters end Sep, Dec, Mar, Jun. The printed year is
        # the calendar year the quarter falls in, so Q1/Q2 open a fiscal year
        # and Q3/Q4 close the one before.
        ENDS = {1: (9, 30), 2: (12, 31), 3: (3, 31), 4: (6, 30)}
        out = {}
        last_year = None
        for c in sorted(quarters):
            q = quarters[c]
            year = years.get(c) or last_year
            if year is None:
                continue
            last_year = year
            month, day = ENDS[q]
            fy = f"FY{year}/{str(year + 1)[2:]}" if q in (1, 2) else f"FY{year - 1}/{str(year)[2:]}"
            out[c] = {"period": f"{year}-{month:02d}-{day:02d}", "fy": fy, "basis": "end"}
        if not out:
            raise ParseError("quarter row found but no year could be resolved")
        return out

    months = {c: month_of(v) for c, v in enumerate(grid[period_row])}
    months = {c: m for c, m in months.items() if m}
    if not months:
        raise ParseError("period row has no parseable months")

    # The year row is usually above the months, but not always: the annual
    # external-debt sheets print "End of | June" first and the years below it.
    # Search upward first, then down.
    year_row = None
    needed = max(1, len(months) // 3)
    search = list(range(period_row - 1, -1, -1)) + list(
        range(period_row + 1, min(len(grid), period_row + 4))
    )
    for i in search:
        hits = sum(1 for c in months if c < len(grid[i]) and year_of(grid[i][c]))
        if hits >= needed:
            year_row = i
            break
    if year_row is None:
        raise ParseError("no year row near the period row")

    # When the years sit below the month labels, the data starts below them.
    data_start = max(period_row, year_row)

    out: dict[int, dict] = {}
    last_year = None
    for c in sorted(months):
        month = months[c]
        year = year_of(grid[year_row][c]) if c < len(grid[year_row]) else None
        if year is None:
            year = last_year
        if year is None:
            continue
        last_year = year

        # Egypt's fiscal year runs 1 July to 30 June, and the archive is laid
        # out that way: months Jul-Dec belong to the printed year, Jan-Jun to
        # the one after. Annual files print the FY end year against "June".
        if frequency == "Annual":
            fy = f"FY{year - 1}/{str(year)[2:]}"
        elif month >= 7:
            fy = f"FY{year}/{str(year + 1)[2:]}"
        else:
            fy = f"FY{year - 1}/{str(year)[2:]}"

        day = LAST_DAY[month]
        if month == 2 and year % 4 == 0 and (year % 100 or year % 400 == 0):
            day = 29
        out[c] = {"period": f"{year}-{month:02d}-{day:02d}", "fy": fy, "basis": "end"}
    out["__data_start__"] = data_start
    return out


def find_label_columns(grid: list[list], first_period_col: int) -> tuple[int | None, int | None]:
    """-> (english column, arabic column), decided by script not position."""
    en = ar = None
    for c in range(min(first_period_col, 4)):
        text = " ".join(
            str(row[c]) for row in grid if c < len(row) and isinstance(row[c], str)
        )
        if not text.strip():
            continue
        if ARABIC.search(text):
            ar = c if ar is None else ar
        else:
            en = c if en is None else en
    return en, ar


def parse_file(path: pathlib.Path, meta: dict) -> tuple[list[tuple], dict[str, dict]]:
    try:
        grid = load_grid(path)
    except Exception:
        grid = load_grid_xls(path)
    if not grid:
        raise ParseError("empty sheet")

    period_row, strategy = find_period_row(grid)
    frequency = meta.get("frequency") or "Monthly"
    periods = build_periods(grid, period_row, frequency, strategy)
    if not periods:
        raise ParseError("no periods resolved")

    data_start = periods.pop("__data_start__", period_row)
    first_col = min(periods)
    en_col, ar_col = find_label_columns(grid, first_col)
    if en_col is None and ar_col is None:
        raise ParseError("no label column found")

    cat = CATEGORY_CODES.get(meta.get("category"), slug(meta.get("category"), 8))
    sub = slug(meta.get("subcategory"), 30)

    observations: list[tuple] = []
    catalog: dict[str, dict] = {}

    for row in grid[data_start + 1 :]:
        label_en = str(row[en_col]).strip() if en_col is not None and en_col < len(row) and row[en_col] else ""
        label_ar = str(row[ar_col]).strip() if ar_col is not None and ar_col < len(row) and row[ar_col] else ""
        label = label_en or label_ar
        if not label:
            continue  # blank separator row

        # Leading whitespace in the label marks a child line in the table.
        raw = str(row[en_col]) if en_col is not None and en_col < len(row) and row[en_col] else str(row[ar_col])
        level = len(raw) - len(raw.lstrip())

        sid = f"EG.XL.{cat}.{sub}.{slug(label_en or label_ar)}"
        values = [
            (periods[c], to_number(row[c]))
            for c in sorted(periods)
            if c < len(row)
        ]
        values = [(p, v) for p, v in values if v is not None]
        if not values:
            continue

        for period, value in values:
            observations.append((sid, period["period"], value))

        catalog.setdefault(
            sid,
            {
                "series_id": sid,
                "title_en": f"{meta.get('subcategory')} - {label_en or label_ar}".strip(),
                "title_ar": label_ar or None,
                "dataset": "excel_archive",
                "family": (meta.get("category") or "").lower().replace(" ", "_"),
                "freq": {"Annual": "A", "Quarterly": "Q", "Monthly": "M"}.get(frequency),
                "unit": None,
                "level": level // 2,
                "period_basis": "end",
                "fiscal_year_basis": "Egypt, 1 July to 30 June",
                "source_file": meta.get("url"),
            },
        )
    return observations, catalog


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--limit", type=int)
    ap.add_argument("--category")
    args = ap.parse_args()

    manifest = json.loads(MANIFEST.read_text(encoding="utf-8"))["files"]
    on_disk = {
        os.path.splitext(os.path.basename(p))[0]: pathlib.Path(p)
        for p in EXCEL.glob("*.xls*")
    }

    todo = [m for m in manifest if m["id"] in on_disk]
    if args.category:
        todo = [m for m in todo if m["category"] == args.category]
    if args.limit:
        todo = todo[: args.limit]
    print(f"{len(todo)} files to parse ({len(manifest)} in manifest, {len(on_disk)} on disk)\n")

    all_obs: list[tuple] = []
    all_cat: dict[str, dict] = {}
    quarantine: list[dict] = []
    ok = 0

    for i, meta in enumerate(todo, 1):
        path = on_disk[meta["id"]]
        try:
            obs, cat = parse_file(path, meta)
        except Exception as exc:  # noqa: BLE001
            quarantine.append(
                {
                    "id": meta["id"],
                    "url": meta["url"],
                    "category": meta.get("category"),
                    "subcategory": meta.get("subcategory"),
                    "frequency": meta.get("frequency"),
                    "reason": f"{type(exc).__name__}: {exc}",
                }
            )
            continue
        if not obs:
            quarantine.append(
                {
                    "id": meta["id"], "url": meta["url"],
                    "category": meta.get("category"),
                    "subcategory": meta.get("subcategory"),
                    "reason": "parsed but produced no observations",
                }
            )
            continue
        all_obs += obs
        all_cat.update(cat)
        ok += 1
        if i % 100 == 0:
            print(f"  [{i}/{len(todo)}] {ok} ok, {len(quarantine)} quarantined")

    # The same series appears in many files (one per fiscal year), so the same
    # (series, period) can be seen twice. Later files supersede earlier ones,
    # which is also how a revision arrives.
    merged: dict[tuple[str, str], float] = {}
    for sid, period, value in all_obs:
        merged[(sid, period)] = value
    rows = sorted((sid, period, value) for (sid, period), value in merged.items())

    OUT_SERIES.parent.mkdir(parents=True, exist_ok=True)
    with OUT_SERIES.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["series_id", "period", "value"])
        w.writerows(rows)

    existing = json.loads(OUT_CATALOG.read_text(encoding="utf-8")) if OUT_CATALOG.exists() else []
    by_id = {c["series_id"]: c for c in existing}
    by_id.update(all_cat)
    OUT_CATALOG.write_text(
        json.dumps(sorted(by_id.values(), key=lambda c: c["series_id"]), indent=1, ensure_ascii=False),
        encoding="utf-8",
    )

    QUARANTINE.write_text(
        json.dumps({"count": len(quarantine), "files": quarantine}, indent=1, ensure_ascii=False),
        encoding="utf-8",
    )

    print(f"\nparsed {ok}/{len(todo)} files")
    print(f"{len(all_cat):,} series, {len(rows):,} observations")
    print(f"catalog now holds {len(by_id):,} series")
    if quarantine:
        reasons = collections.Counter(q["reason"].split(":")[0] for q in quarantine)
        print(f"\nquarantined {len(quarantine)} files -> catalog/excel_quarantine.json")
        for reason, n in reasons.most_common(8):
            print(f"   {reason}: {n}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
